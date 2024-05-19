import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import os

# 파일 경로 리스트
file_list = ["C:/Users/iks12/Desktop/11/22", "C:/Users/iks12/Desktop/11/33", "C:/Users/iks12/Desktop/11/44"]

# 파일 경로 리스트를 데이터프레임으로 변환
df = pd.DataFrame(file_list, columns=["File Path"])

# 엑셀 파일로 저장
excel_path = "C:/Users/iks12/Desktop/11/123.xlsx"  # 확장자 추가
df.to_excel(excel_path, index=False)

# 저장한 엑셀 파일 열기
wb = load_workbook(excel_path)
ws = wb.active

# 하이퍼링크 설정
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=ws.max_row):
    for cell in row:
        file_path = cell.value
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=file_path, tooltip="Open file")
        cell.font = Font(color="0000FF", underline="single")

# 변경된 엑셀 파일 저장
wb.save(excel_path)

# 엑셀 파일을 자동으로 열기 (윈도우)
os.startfile(excel_path)

print("파일 경로가 하이퍼링크로 저장된 엑셀 파일이 생성되고 열렸습니다.")
